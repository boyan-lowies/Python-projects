
import openpyxl as op
import pytesseract
from pytesseract import Output
import pyautogui as gui
import random
import re
import time
from PIL import ImageGrab
from difflib import SequenceMatcher
import ctypes
import requests

wb = op.load_workbook('C:\\Users\\lowie\\OneDrive\\Guild Wars 2.xlsx', data_only=True)
sheetTP = wb['TP current']
sheetD = wb['Data']
sheetN = wb['Next']

pytesseract.pytesseract.tesseract_cmd = r'C:\\Program Files\\Tesseract-OCR\\tesseract.exe'

Searchbox = [67, -221, 341, -195]
Itembox = [67, -181, 367, 591]
Countbox = [589,392,598,403]
Craftbox = [627,384,767,406]
pixelsearch = [939, 390, 949, 400]
statics = [77256, 19790, 76839, 46747]

disiplin = ['Armorsmith', 'Leatherworker', 'Tailor']
# 'Huntsman', 'Weaponsmith','Artificer'

#------------------------#
#      Subroutines       #
#------------------------#
 

def search_value_in_column(ws, search_string, column):
    for row in range(1, ws.max_row + 1):
        coordinate = "{}{}".format(column, row)
        if ws[coordinate].value == search_string:
            return row
    return

def user_input(id):
   print("Item id", id, "not found, print name here:")
   name = input()
   return name

def move_cords(location, offsetx, offsety, randofset):
    location = [location[0] + zeropoint[0],location[1] + zeropoint[1], location[2] - location[0], location[3] - location[1]]
    location = [location[0]+ random.randint(0+randofset,location[2]-randofset) + offsetx, location[1] + random.randint(0+randofset,location[3] - randofset) + offsety]
    gui.moveTo(location[0], location[1], 1, gui.easeInOutQuad) 
    gui.click(location[0] ,location[1]) 

def delete_line():
    gui.moveTo(zeropoint[0] + 334, zeropoint[1] - 207, 1, gui.easeInOutQuad)
    time.sleep(0.5)
    gui.dragTo(zeropoint[0] + 51, zeropoint[1] - 208, 1, gui.easeInOutQuad)

def move_slider():
    gui.moveTo(zeropoint[0] + 373, zeropoint[1] - 150, 1, gui.easeInOutQuad)
    time.sleep(0.5)
    gui.dragTo(zeropoint[0] + 373, zeropoint[1] +555, 1, gui.easeInOutQuad)

def wait_done():
    search = [pixelsearch[0] + zeropoint[0], pixelsearch[1] + zeropoint[1], pixelsearch[2] + zeropoint[0], pixelsearch[3] + zeropoint[1]]
    while 1:
        time.sleep(1)
        img = ImageGrab.grab(search)
        img.save('plz.png')
        img = img.load()
        colour = img[1,1]
        if colour[1] > 150:
            break


#------------------------#
#      Drill list        #
#------------------------#

sub_items = ["AF","AH","AJ","AL","AN"]
sub_count = ["AG","AI","AK","AM"]

def drill_list(row, buy, amount, itemc_count):
    item = []
    idrow = []
    count = []
    name = []
    item_b = []
    count_b = []
    item_craftcount = []

    for x in range(0, len(row)):
        adress = sub_items[0] + str(row[x])
        z=0   
        while sheetD[adress].value != None:
            y = sheetD[adress].value
            if y not in item:               
                item.append(y) 
                count.append((sheetD[sub_count[z] + str(row[x])].value)*amount[x]*(1/itemc_count[x]))
            else:
                index = item.index(y)
                count[index] +=  (sheetD[sub_count[z] + str(row[x])].value)*amount[x]*(1/itemc_count[x])
            z += 1
            adress = sub_items[z] + str(row[x])

    for x in item:
        if x not in buy:
            index = item.index(x)
            temp = search_value_in_column(sheetD,x,"AC")
            idrow.append(temp)
            item_b.append(item[index])
            count_b.append(count[index])
            item_craftcount.append(sheetD['AD' + str(temp)].value)
            try:
                name.append(sheetD["B"+str(search_value_in_column(sheetD,x,"A"))].value)
            except:
                req = "https://api.guildwars2.com/v2/items/" + str(x)
                result = requests.get(req)
                result = result.json()
                name.append(result['name'])
           


    return item_b, count_b, idrow, name, item_craftcount


#------------------------#
#     Move to item id    #
#------------------------#

def move_id(name_r, randofset, count_r):
    slid = 0
    while 1:
        certainty =[]
        itemnames = []
        name = name_r.replace(" ","")
        name = name[:22]
        search = [Itembox[0] + zeropoint[0], Itembox[1] + zeropoint[1], Itembox[2] + zeropoint[0], Itembox[3] + zeropoint[1]]
        img = ImageGrab.grab(search)
        img.save('test.png')
        d = pytesseract.image_to_data(img, output_type=Output.DICT,lang='eng', config='--psm 4 --oem 1 -c tessedit_char_whitelist=qwertyuiopQWERTYUIOPasdfghjklASDFGHJKLzxcvbnmZXCVBNM(1234567890)')
        for y in range(len(d['text'])):
            test = re.sub("\(.*?\)","", d['text'][y])
            test = test[:22]
            itemnames.append(test)
            certainty.append(SequenceMatcher(None, name, test).ratio())
        count = max(certainty) 
        x = certainty.index(count)
        
        if count > 0.90:
            break
        if count > 0.75:
            string = "Are\n "+str(name_r)+" \nand\n "+str(itemnames[x])+" \nthe same item?"
            var = ctypes.windll.user32.MessageBoxW(0, string, "", 4)
            if var == 6:
                break

        if slid == 1:
            string = "Couldnt find "+ str(name_r) + ", get reciepe or craft "+ str(count_r) + "items"
            ctypes.windll.user32.MessageBoxW(0, string, "", 4)
            return

        move_slider()
        slid = 1

    location = [d['left'][x]+ random.randint(0+randofset,d['width'][x]-randofset) + search[0], d['top'][x] + random.randint(0+randofset,d['height'][x] - randofset) + search[1]]
    gui.moveTo(location[0], location[1], 1, gui.easeInOutQuad) 
    gui.click(location[0] ,location[1]) 
    return 

#------------------------#
#         Main           #
#------------------------#

buy_id = []
N=2
adress = "O2"

while sheetN[adress].value is not None: 
    buy_id.append(sheetN[adress].value)
    N += 1
    adress = "O" + str(N)

buy_id= buy_id + statics

items_b=[]
itemid_b=[]
itemc_b=[]
itemc_row_b=[]
itemc_count_b=[]

N=2
adress= "B2" 

while sheetN[adress].value is not None:
    items_b.append(sheetN["B" + str(N)].value)
    x=len(items_b)-1
    itemid_b.append(sheetD["A" + str(search_value_in_column(sheetD,items_b[x],"B"))].value)
    itemc_row_b.append(search_value_in_column(sheetD,itemid_b[x],"AC"))
    itemc_b.append(sheetD["AB" + str(itemc_row_b[x])].value)
    itemc_count_b.append(sheetD["AD" + str(itemc_row_b[x])].value)

    N += 1
    adress= "B" + str(N)  

print(itemc_b)

for q in disiplin:

    items=[]
    itemid=[]
    itemc=[]
    itemc_count = []
    item_count=[]
    itemc_row=[]
    list = []
    z = 0

    for y in range(len(items_b)):
        if itemc_b[y] != q:
            continue
        
        items.append(items_b[y])
        itemid.append(itemid_b[y])
        itemc.append(itemc_b[y])
        item_count.append(1)
        itemc_row.append(itemc_row_b[y])
        itemc_count.append(itemc_count_b[y])

    list.append(drill_list(itemc_row, buy_id, item_count, itemc_count))

    while len(list[z][0]) != 0:
        list.append(drill_list(list[z][2], buy_id, list[z][1], list[z][4]))
        z += 1
    
    zeropoint = gui.locateOnScreen('C:\\Users\\lowie\\wardrobe.jpg', confidence=0.9)
    print(zeropoint)

    for p in range(len(list)-1,-1, -1):
        for x in range(len(list[p][3])):
            delete_line()
            time.sleep(0.5) 
            gui.write(list[p][3][x], interval=0.1)
            print(list[p][3][x], list[p][1][x])
            move_id(list[p][3][x],3,list[p][1][x])
            time.sleep(0.5)
            move_cords(Countbox,0,0,1)
            time.sleep(0.5)
            gui.write(str(list[p][1][x]), interval=0.1)
            time.sleep(0.5)
            move_cords(Craftbox,0,0,5)
            wait_done()
    
    for x in range(len(items)):
        delete_line()
        time.sleep(0.5) 
        gui.write(items[x], interval=0.1)
        move_id(items[x],3,1)
        time.sleep(0.5)
        move_cords(Countbox,0,0,0)
        time.sleep(0.5)
        gui.write(str(1), interval=0.1)
        time.sleep(0.5)
        move_cords(Craftbox,0,0,0)
        wait_done()

    nextthing = disiplin.index(q)
    nextthing = disiplin[nextthing+1]
    string = "Change to character with " + str(nextthing) + "profession"
    ctypes.windll.user32.MessageBoxW(0, string, "", 4)
print("Done")
   