
from matplotlib.pyplot import pause
import openpyxl as op

wb = op.load_workbook('C:\\Users\\lowie\\Desktop\\pytest.xlsx', data_only=True)
sheetTP = wb['TP current']
sheetD = wb['Data']
sheetN = wb['Next']

statics = [77256, 19790, 76839]



#------------------------#
#      Subroutines       #
#------------------------#
 

def search_value_in_column(ws, search_string, column):
    for row in range(1, ws.max_row + 1):
        coordinate = "{}{}".format(column, row)
        if ws[coordinate].value == search_string:
            return row
    return

def user_input(id):
   print("Item id", id, "not found, print name here:")
   name = input()
   return name
#------------------------#
#      Drill list        #
#------------------------#

sub_items = ["AF","AH","AJ","AL","AN"]
sub_count = ["AG","AI","AK","AM"]


def drill_list(row, buy):
    item = []
    idrow = []
    count = []
    name = []
    item_b = []
    count_b = []

    for x in range(0, len(row)):
        adress = sub_items[0] + str(row[x])
        z=0   
        while sheetD[adress].value != None:
            y = sheetD[adress].value
            if y not in item:               
                item.append(y) 
                count.append(sheetD[sub_count[z] + str(row[x])].value)
            else:
                index = item.index(y)
                count[index] = count[index] + sheetD[sub_count[z] + str(row[x])].value
            z += 1
            adress = sub_items[z] + str(row[x])

    for x in item:
        if x not in buy:
            index = item.index(x)
            idrow.append(search_value_in_column(sheetD,str(x),"AC"))
            item_b.append(item[index])
            count_b.append(count[index])
            try:
                name.append(sheetD["B"+str(search_value_in_column(sheetD,x,"A"))].value)
            except:
                name.append(user_input(x))
            print(name)


    return item_b, count_b, idrow, name


#------------------------#
#         Main           #
#------------------------#



items=[]
itemid=[]
itemc=[]
itemc_row=[]

N=2
adress= "B2"  


while sheetN[adress].value is not None:
    items.append(sheetN["B" + str(N)].value)
    x=len(items)-1
    itemid.append(sheetD["A" + str(search_value_in_column(sheetD,items[x],"B"))].value)
    itemc_row.append(search_value_in_column(sheetD,str(itemid[x]),"AC"))
    itemc.append(sheetD["AB" + str(itemc_row[x])].value)
    N += 1
    adress= "B" + str(N)  


buy_id = []
N=2
adress = "O2"

while sheetN[adress].value is not None: 
    buy_id.append(sheetN[adress].value)
    N += 1
    adress = "O" + str(N)

buy_id= buy_id + statics

list_1 = drill_list(itemc_row, buy_id)
print(list_1)
list_2 = drill_list(list_1[2], buy_id)
print(list_2)
list_3 = drill_list(list_2[2], buy_id)
print(list_3)
list_4 = drill_list(list_3[2], buy_id)
print(list_4)
list_5 = drill_list(list_4[2], buy_id)
print(list_5)
